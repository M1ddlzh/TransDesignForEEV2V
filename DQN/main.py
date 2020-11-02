import datetime
import os
import matplotlib.pyplot as plt
import numpy as np
import tensorflow as tf
from openpyxl import Workbook
from tqdm import tqdm

from env_twousers_dqn import EnvTwoUsers
from RL_brain import DeepQNetwork


ep_len = 30     # T = 30

def testing_while_training(env, RL):
    r_test = 0
    info_test = {'ty1_succ_rate_1': 0, 'ty1_succ_rate_2': 0, 
        'Q_len_1': 0, 'Q_len_2': 0, 
        'energy_effi_1': 0, 'energy_effi_2': 0,  
        'avg_rate': 0, 'avg_power':0,}
    eps = 100
    for i in range(eps):
        observation = env.reset()
        while True:
            action, _= RL.choose_action_test(observation)
            observation_, reward, done, info = env.step(action)
            r_test += reward
            if done:
                for k, v in info.items():
                    if k in info_test.keys():
                        info_test[k] += v
                    else:
                        info_test[k] = v
                break
            observation = observation_
    r_test /= eps
    for k, _ in info_test.items():
        info_test[k] /= eps

    return r_test, info_test

def run(env, RL, episodes, path):
    slot = 0
    best_ep, best_reward = -1, -1
    training_res = []
    for episode in tqdm(range(1, episodes + 1), ncols=60):
        # initial observation
        observation = env.reset()
        while True:
            # RL choose action based on observation
            action, _= RL.choose_action(observation)
            # RL take action and get next observation and reward
            observation_, reward, done, info = env.step(action)
            RL.store_transition(observation, action, reward, 
                done, observation_)
            slot += 1
            if slot > 200:
                RL.learn()
            if done:
                # if episode % 50 == 0:
                #     print('\n' + str(episode) + '/' + str(episodes))
                #     print(info)
                if episode % 50 == 0: # or episode < 2000:
                    env_test = EnvTwoUsers(ep_len)
                    # env_test.seed(0)
                    r_test, info_test = testing_while_training(env_test, RL)
                    print(info_test)
                    training_res.append([
                        (info_test['energy_effi_1'] 
                        + info_test['energy_effi_2']) / 2,
                        (info_test['ty1_succ_rate_1'] 
                        + info_test['ty1_succ_rate_2']) / 2,
                        (info_test['Q_len_1'] + info_test['Q_len_1']) / 2,
                        r_test])
                    if best_ep == -1 or best_reward < r_test:
                        best_reward = r_test
                        best_ep = episode
                        RL.save_fn(path)
                break
            observation = observation_
    # end of training
    training_res = np.array(training_res)
    wb = Workbook()
    ws = wb.active
    ws.title = 'training result'
    ws['A1'] = 'testing num'
    ws['B1'] = 'energy efficiency'
    ws['C1'] = 'type 1 success rate'
    ws['D1'] = 'type 2 q length'
    ws['E1'] = 'return'
    for i in range(training_res.shape[0]):
        ws.cell(i+2, 1).value = i + 1
        ws.cell(i+2, 2).value = training_res[i, 0]
        ws.cell(i+2, 3).value = training_res[i, 1]
        ws.cell(i+2, 4).value = training_res[i, 2]
        ws.cell(i+2, 5).value = training_res[i, 3]
    wb.save("training_directly_slot30.xlsx")
    print('training over')

    # testing
    env_test = EnvTwoUsers(ep_len)
    res = {'ty1_succ_rate_1': 0, 'ty1_succ_rate_2': 0, 
        'Q_len_1': 0, 'Q_len_2': 0, 
        'energy_effi_1': 0, 'energy_effi_2': 0,  
        'avg_rate': 0, 'avg_power':0,}
    smaller_than_phi = 0
    larger_than_qmax = 0
    RL.model.load_weights(path)
    print('reload model!')
    eps = 10000
    for ep in tqdm(range(eps), ncols=60):
        s = env_test.reset()
        while True:
            action, _ = RL.choose_action_test(s)
            s_, reward, done, info = env_test.step(action)
            if done:
                for k, v in info.items():
                    if k in res.keys():
                        res[k] += v
                    else:
                        res[k] = v
                if info['ty1_succ_rate_1'] < 1 or info['ty1_succ_rate_2'] < 1:
                    smaller_than_phi += 1
                if info['Q_len_1'] > env_test.Q_max \
                    or info['Q_len_2'] > env_test.Q_max:
                    larger_than_qmax += 1
                break
            s = s_
    for k, v in res.items():
        print(k, v / eps)
    print('\nsmaller_than_phi:', smaller_than_phi)
    print('larger_than_qmax:', larger_than_qmax)


if __name__ == "__main__":
    env = EnvTwoUsers(ep_len)
    RL = DeepQNetwork(env.n_actions, env.n_features,
        learning_rate=0.0001,
        reward_decay=1.,
        e_greedy_final=0.1,
        replace_target_iter=500,
        memory_size=20000,
        batch_size=128, 
        e_greedy_decrement=0.0001, 
        num_layers=4, 
        num_units=32)
    path = './dqn_model/dqn_checkpoint'
    if not os.path.exists(path):
        os.makedirs(path)
    episodes = 100000
    run(env, RL, episodes, path)


