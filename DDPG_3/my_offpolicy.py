import time
import tqdm
from torch.utils.tensorboard import SummaryWriter
from typing import Dict, List, Union, Callable, Optional

from tianshou.policy import BasePolicy
from tianshou.utils import tqdm_config, MovAvg
from tianshou.trainer import test_episode, gather_info

# change
from my_ddpg_collector import Collector
from env_threeusers_ddpg import EnvThreeUsers

import numpy as np
from openpyxl import Workbook


def offpolicy_trainer(
        policy: BasePolicy,
        train_collector: Collector,
        test_collector: Collector,
        max_epoch: int,
        step_per_epoch: int,
        collect_per_step: int,
        episode_per_test: Union[int, List[int]],
        batch_size: int,
        update_per_step: int = 1,
        train_fn: Optional[Callable[[int], None]] = None,
        test_fn: Optional[Callable[[int], None]] = None,
        stop_fn: Optional[Callable[[float], bool]] = None,
        save_fn: Optional[Callable[[BasePolicy], None]] = None,
        log_fn: Optional[Callable[[dict], None]] = None,
        writer: Optional[SummaryWriter] = None,
        log_interval: int = 1,
        verbose: bool = True,
        # test_in_train: bool = True,
        test_in_train: bool = False,
) -> Dict[str, Union[float, str]]:
    """A wrapper for off-policy trainer procedure.

    :param policy: an instance of the :class:`~tianshou.policy.BasePolicy`
        class.
    :param train_collector: the collector used for training.
    :type train_collector: :class:`~tianshou.data.Collector`
    :param test_collector: the collector used for testing.
    :type test_collector: :class:`~tianshou.data.Collector`
    :param int max_epoch: the maximum of epochs for training. The training
        process might be finished before reaching the ``max_epoch``.
    :param int step_per_epoch: the number of step for updating policy network
        in one epoch.
    :param int collect_per_step: the number of frames the collector would
        collect before the network update. In other words, collect some frames
        and do some policy network update.
    :param episode_per_test: the number of episodes for one policy evaluation.
    :param int batch_size: the batch size of sample data, which is going to
        feed in the policy network.
    :param int update_per_step: the number of times the policy network would
        be updated after frames be collected. In other words, collect some
        frames and do some policy network update.
    :param function train_fn: a function receives the current number of epoch
        index and performs some operations at the beginning of training in this
        epoch.
    :param function test_fn: a function receives the current number of epoch
        index and performs some operations at the beginning of testing in this
        epoch.
    :param function save_fn: a function for saving policy when the undiscounted
        average mean reward in evaluation phase gets better.
    :param function stop_fn: a function receives the average undiscounted
        returns of the testing result, return a boolean which indicates whether
        reaching the goal.
    :param function log_fn: a function receives env info for logging.
    :param torch.utils.tensorboard.SummaryWriter writer: a TensorBoard
        SummaryWriter.
    :param int log_interval: the log interval of the writer.
    :param bool verbose: whether to print the information.
    :param bool test_in_train: whether to test in the training phase.

    :return: See :func:`~tianshou.trainer.gather_info`.
    """
    global_step = 0
    best_epoch, best_reward = -1, -1
    stat = {}
    start_time = time.time()
    test_in_train = test_in_train and train_collector.policy == policy
    # change
    training_res = []
    for epoch in range(1, 1 + max_epoch):
        # train
        policy.train()
        if train_fn:
            train_fn(epoch)
        with tqdm.tqdm(total=step_per_epoch, desc=f'Epoch #{epoch}',
                       **tqdm_config) as t:
            while t.n < t.total:
                result = train_collector.collect(n_step=collect_per_step,
                                                 log_fn=log_fn)
                # data = {}
                # if test_in_train and stop_fn and stop_fn(result['rew']):
                #     test_result = test_episode(
                #         policy, test_collector, test_fn,
                #         epoch, episode_per_test)
                #     if stop_fn and stop_fn(test_result['rew']):
                #         if save_fn:
                #             save_fn(policy)
                #         for k in result.keys():
                #             data[k] = f'{result[k]:.2f}'
                #         t.set_postfix(**data)
                #         return gather_info(
                #             start_time, train_collector, test_collector,
                #             test_result['rew'])
                #     else:
                #         policy.train()
                #         if train_fn:
                #             train_fn(epoch)
                for i in range(update_per_step * min(
                        result['n/st'] // collect_per_step, t.total - t.n)):
                    global_step += 1
                    losses = policy.learn(train_collector.sample(batch_size))
                    # for k in result.keys():
                    #     data[k] = f'{result[k]:.2f}'
                    #     if writer and global_step % log_interval == 0:
                    #         writer.add_scalar(
                    #             k, result[k], global_step=global_step)
                    # for k in losses.keys():
                    #     if stat.get(k) is None:
                    #         stat[k] = MovAvg()
                    #     stat[k].add(losses[k])
                    #     data[k] = f'{stat[k].get():.6f}'
                    #     if writer and global_step % log_interval == 0:
                    #         writer.add_scalar(
                    #             k, stat[k].get(), global_step=global_step)
                    t.update(1)
                    # change
                    # t.set_postfix(**data)
            if t.n <= t.total:
                t.update()
        # test
        # change
        if epoch % 50 == 0: # or epoch < 2000:
            env = EnvThreeUsers(step_per_epoch)
            # env.seed(0)
            policy.train(False)
            collector = Collector(policy, env)
            ep = 100
            result = collector.collect(n_episode=ep)
            # result = test_episode(
            #     policy, test_collector, test_fn, epoch, episode_per_test)
            if best_epoch == -1 or best_reward < result['rew']:
                best_reward = result['rew']
                best_epoch = epoch
                if save_fn:
                    save_fn(policy)
            # print(result)
            if verbose:
                # change
                print(f'Epoch #{epoch}: test_reward: {result["rew"]:.6f}, '
                    f'best_reward: {best_reward:.6f} in #{best_epoch}\n'
                    f'ty1_succ_rate_1: {result["ty1s_1"][0]/ep:.4f}, '
                    f'ty1_succ_rate_2: {result["ty1s_2"][0]/ep:.4f}, ', 
                    f'ty1_succ_rate_3: {result["ty1s_3"][0]/ep:.4f}, \n', 
                    f'Q_len_1: {result["ql_1"][0]/ep:.4f},' 
                    f'Q_len_2: {result["ql_2"][0]/ep:.4f},',
                    f'Q_len_3: {result["ql_3"][0]/ep:.4f}, \n',
                    f'energy_effi_1: {result["ee_1"][0]/ep:.4f},'
                    f'energy_effi_2: {result["ee_2"][0]/ep:.4f},', 
                    f'energy_effi_3: {result["ee_3"][0]/ep:.4f}\n', 
                    f'avg_rate: {result["avg_r"]/ep:.4f}, '
                    f'avg_power: {result["avg_p"]/ep:.4f} dBm\n')
            # change
            training_res.append([(result["ee_1"][0]/ep + result["ee_2"][0]/ep
                + result["ee_3"][0]/ep) / 3, 
                (result["ty1s_1"][0]/ep + result["ty1s_2"][0]/ep
                + result["ty1s_3"][0]/ep) / 3,
                (result["ql_1"][0]/ep + result["ql_2"][0]/ep
                + result["ql_2"][0]/ep) / 3,
                result["rew"]])
        if stop_fn and stop_fn(best_reward):
            break
    # change
    training_res = np.array(training_res)
    wb = Workbook()
    ws = wb.active
    ws.title = 'training result'
    ws['A1'] = 'testing num'
    ws['B1'] = 'energy efficiency'
    ws['C1'] = 'type 1 success rate'
    ws['D1'] = 'type 2 q length'
    ws['E1'] = 'return'
    for i in range(training_res.shape[0]):
        ws.cell(i+2, 1).value = i + 1
        ws.cell(i+2, 2).value = training_res[i, 0]
        ws.cell(i+2, 3).value = training_res[i, 1]
        ws.cell(i+2, 4).value = training_res[i, 2]
        ws.cell(i+2, 5).value = training_res[i, 3]
    wb.save("directly_training_slot" + str(step_per_epoch) + ".xlsx")
    test_collector.collect_time = -1
    return gather_info(
        start_time, train_collector, test_collector, best_reward)
